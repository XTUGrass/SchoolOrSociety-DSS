from pybrain.datasets import ClassificationDataSet
from pybrain.utilities import percentError
from pybrain.tools.shortcuts import buildNetwork
from pybrain.supervised.trainers import BackpropTrainer
from pybrain.structure.modules import SoftmaxLayer
from pybrain.tools.validation import Validator

from openpyxl import load_workbook

import matplotlib.pyplot as plt

class model1:
    wb = load_workbook('e:\\python/django/demo_schoolOrSociety/dss_model/test.xlsx')
    ws = wb.active

    alldata = ClassificationDataSet(37, 1, nb_classes=2)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=39):
        temp_array = []
        for cell in row:
            temp_array.append(int(cell.value))
        alldata.addSample((temp_array[1], temp_array[2], temp_array[3], temp_array[4], temp_array[5],
                           temp_array[6], temp_array[7], temp_array[8], temp_array[9], temp_array[10],
                           temp_array[11], temp_array[12], temp_array[13], temp_array[14], temp_array[15],
                           temp_array[16], temp_array[17], temp_array[18], temp_array[19], temp_array[20],
                           temp_array[21], temp_array[22], temp_array[23], temp_array[24], temp_array[25],
                           temp_array[26], temp_array[27], temp_array[28], temp_array[29], temp_array[30],
                           temp_array[31], temp_array[32], temp_array[33], temp_array[34], temp_array[35],
                           temp_array[36], temp_array[37]), temp_array[0])

    tstdata_temp, trndata_temp = alldata.splitWithProportion(0.25)

    tstdata = ClassificationDataSet(37, 1, nb_classes=2)
    for n in xrange(0, tstdata_temp.getLength()):
        tstdata.addSample(tstdata_temp.getSample(n)[0], tstdata_temp.getSample(n)[1])

    trndata=ClassificationDataSet(37, 1, nb_classes=2)
    for n in xrange(0, trndata_temp.getLength()):
        trndata.addSample(trndata_temp.getSample(n)[0], trndata_temp.getSample(n)[1])

    trndata._convertToOneOfMany()
    tstdata._convertToOneOfMany()

    fnn = buildNetwork(trndata.indim, 6, trndata.outdim, outclass=SoftmaxLayer)
    trainer = BackpropTrainer(fnn, dataset=trndata, momentum=0.1, verbose=True, weightdecay=0.01)

    for i in range(20):
        trainer.trainEpochs(10)


    def getResult(self, paraArray):
        result = self.fnn.activate(paraArray)
        if result[0] > result[1]:
            return 1  # 1 for kaoyan
        else:
            return 2  # 2 for work
